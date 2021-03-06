import os
import openpyxl
import csv
from split import split
from sendfiles import send_files
from parse import parse_row
import unicodecsv

columns_to_delete = [8, 13, 14, 15, 18, 21, 24, 26, 30, 55, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 73, 75, 87, 88,
                     89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105, 106, 112, 113, 116, 120,
                     121, 122, 123, 127, 131, 132, 133, 134, 135, 136, 138, 139, 140, 141, 142, 143, 144, 147]

last_id = 2


def format_special_characters(value):
    value = value.replace(';', ',')
    value = value.replace('\\', '')

    return value


def create_excel_from_csv(file_path, count):
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(file_path, encoding='utf-8') as csv_file:
        reader = csv.reader(csv_file, delimiter=';')
        for row in reader:
            if row[107] not in ('4', '5', 'CLASSI_FIN'):
                continue
            else:
                row = parse_row(row)
                new_row = []
                for i in range(0, len(row)):
                    print(row)
                    if i not in columns_to_delete:
                        row[i] = format_special_characters(row[i])
                        new_row.append(row[i])
                ws.append(new_row)

    excel_file_name = (cwd + str(count) + '_temp_excelfile.xlsx').replace('\\', '/')
    wb.save(excel_file_name)

    return excel_file_name


def open_new_excel_file_and_insert_firstrow(new_wb_sheet):
    print("criar coluna id")
    print(new_wb_sheet)
    new_wb_sheet.insert_cols(1)
    new_wb_sheet['A1'] = '_ID'


def add_ids_to_srag_records(new_wb_sheet, last_id):
    print("adiciona id")
    print(new_wb_sheet)

    for id in range(2, (new_wb_sheet.max_row + 1)):
        cell_id = 'A' + str(id)
        new_wb_sheet[str(cell_id)] = last_id - 1
        last_id += 1

    return last_id


def create_new_formatted_csv_file(new_wb_sheet, count):
    print("criando novo csv")
    print(new_wb_sheet)
    new_csv_filename = str(os.getcwd().replace('\\', '/')) + '/ready_files/' + str(count) + '_finalcsv.csv'
    csvfile = open(new_csv_filename, 'w', encoding='cp860', newline='')
    wr = csv.writer(csvfile, quoting=csv.QUOTE_ALL, delimiter=";")
    for row in new_wb_sheet.iter_rows():
        lrow = []
        for cell in row:
            x = str(cell.value)
            lrow.append(x)
        wr.writerow(lrow)
    print(" ... done")
    csvfile.close()




    # for row in new_wb_sheet.rows:
    #     l = list(row)
    #     for i in range(len(l)):
    #         if i == len(l) - 1:
    #             if str(l[i].value) == 'None':
    #                 csv.write('' + ';')
    #             else:
    #                 csv.write(str(l[i].value).encode().decode())
    #         else:
    #             if str(l[i].value) == 'None':
    #                 csv.write('' + ';')
    #             else:
    #                 csv.write(str(l[i].value).encode().decode() + ';')
    #     csv.write('\n')
    #
    # csv.close()

def sort_list(files_list):
    sorted_list = []
    print(files_list)
    for i in range(0, len(files_list)):
        filename = cwd + str(i + 1) + '_temp_csvfile.csv'
        print(filename)
        sorted_list.append(filename)

    return sorted_list


def run_on_exit():
    for file_to_delete in os.listdir(cwd):
        os.remove(cwd + file_to_delete)


### Start of main.py execution ###

with open('file_to_upload/dummydata3.csv', encoding='utf-8') as csv_big_file:
    split(csv_big_file)

get_all_files = []
cwd = os.getcwd() + '\\temp_files\\'
for file in os.listdir(cwd):
    if file.endswith("temp_csvfile.csv"):
        get_all_files.append(os.path.join(cwd, file).replace('\\', '/'))

all_files_list = sort_list(get_all_files)

print(all_files_list)

count = 1

for file in all_files_list:
    print("começou execução arquivo " + str(count))
    excel_new_file = create_excel_from_csv(file, count)

    new_wb = openpyxl.load_workbook(excel_new_file)

    new_wb_sheet = new_wb.worksheets[0]

    open_new_excel_file_and_insert_firstrow(new_wb_sheet)
    last_id = add_ids_to_srag_records(new_wb_sheet, last_id)
    new_wb.save(excel_new_file)
    create_new_formatted_csv_file(new_wb_sheet, count)

    new_wb.close()
    print("terminou execução arquivo " + str(count))
    count += 1

# run_on_exit()
send_files()
